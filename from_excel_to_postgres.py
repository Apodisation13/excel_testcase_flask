# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

from models import *
from db import *


# Load in the workbook
wb = load_workbook('./test_data.xlsx')

sheet = wb['данные']

# Print the sheet title
print(sheet.title)

max_row = sheet.max_row
max_column = sheet.max_column
print(max_row, max_column)

for row in range(2, max_row + 1):  # с 2го, ибо 1й ведь шапка

    # print(sheet.cell(row=row, column=column).value)

    result = sheet.cell(row=row, column=4).value * sheet.cell(row=row, column=5).value

    new = TestExcel(
        mnn=sheet.cell(row=row, column=1).value,
        trade_name=sheet.cell(row=row, column=2).value,
        form=sheet.cell(row=row, column=3).value,
        quantity=sheet.cell(row=row, column=4).value,
        price=sheet.cell(row=row, column=5).value,
        result=result
    )
    session.add(new)

session.commit()


query = session.query(TestExcel).distinct('mnn')
print(len(query.all()))

q_list = []
price_list = []
result_list = []
i = 1
for distinct_mnn in query:

    q_pr = session.query(TestExcel).filter(TestExcel.mnn == distinct_mnn.mnn)

    # print(q_pr.all())

    quantity = []
    price = []
    result = []
    form = ''
    name = ''

    for each in q_pr.all():
        quantity.append(each.quantity)
        price.append(each.price)
        result.append(each.result)
        form += each.form + '<br>'
        name += each.trade_name + '<br>'
    q_list.append(sum(quantity))
    price_list.append(sum(price))
    result_list.append(sum(result))

    # print(i, distinct_mnn, name, form, q_list, price_list, result_list)
    i += 1

    new = AggrExcel(
        mnn=distinct_mnn.mnn,
        trade_name=name,
        form=form,
        quantity=sum(quantity),
        price=sum(price),
        result=sum(result)
    )
    session.add(new)
session.commit()

sum_quantity = sum(q_list)
sum_price = sum(price_list)
sum_result = sum(result_list)
print(sum_quantity, sum_price, sum_result)
